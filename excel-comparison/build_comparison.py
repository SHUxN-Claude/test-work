# -*- coding: utf-8 -*-
"""前回/今回シートから比較シートを生成するスクリプト。
VBAマクロ(比較シート作成.bas)と同じロジック。
"""
import re
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/root/.claude/uploads/ff77b9b4-9504-506f-b7be-aa7ff9919ea5/7b5d4609-_202601_____068.500095.000_______.xlsx"
OUT = "/tmp/claude-0/-home-user-test-work/ff77b9b4-9504-506f-b7be-aa7ff9919ea5/scratchpad/軌道検査_前回今回比較.xlsx"

TH_LOW, TH_HIGH = 10.0, 15.0

# (項目名, 測定値ヘッダ, 今回目標値Fヘッダ or None)
ITEMS = [
    ("高低（左）", "高低（左）", "高低（左）目標値F"),
    ("高低（右）", "高低（右）", "高低（右）目標値F"),
    ("通り（左）", "通り（左）", "通り（左）目標値F"),
    ("通り（右）", "通り（右）", "通り（右）目標値F"),
    ("水準実測値", "水準実測値", None),
    ("水準", "水準", "水準目標値F"),
    ("平面性", "平面性", "平面性目標値F"),
    ("軌間実測値", "軌間実測値", "軌間実測値目標値F"),
    ("軌間", "軌間", "軌間目標F"),
]
KEY_HEADERS = ["線名コード", "線別コード", "番線番号", "本準側コード", "キロ程"]

FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")
FILL_YELLOW = PatternFill("solid", fgColor="FFF2A8")   # 今回目標値F=1
FILL_ORANGE = PatternFill("solid", fgColor="FFC000")   # 10mm以上15mm未満
FILL_RED    = PatternFill("solid", fgColor="FF6B6B")   # 15mm以上
FILL_WARN   = PatternFill("solid", fgColor="FF6B6B")
FILL_OK     = PatternFill("solid", fgColor="C6EFCE")
BOLD = Font(bold=True)

NUM_RE = re.compile(r"^-?\d+(\.\d+)?$")


def to_num(v):
    """数値化できれば int/float、できなければ None(空) または元の文字列。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "":
        return None
    if NUM_RE.match(s):
        f = float(s)
        if "." not in s:
            return int(f)
        return f
    return v  # 文字列のまま


def read_sheet(ws):
    """ヘッダ辞書と行リスト(生値)を返す。"""
    rows = list(ws.iter_rows(values_only=True))
    header = [("" if h is None else str(h).strip()) for h in rows[0]]
    hidx = {}
    for i, h in enumerate(header):
        if h and h not in hidx:
            hidx[h] = i
    data = [r for r in rows[1:] if any(c not in (None, "") for c in r)]
    return header, hidx, data


def fnum(v):
    """厳密な float 化。数値でなければ None。"""
    n = to_num(v)
    return float(n) if isinstance(n, (int, float)) else None


def build():
    print("reading source ...")
    swb = openpyxl.load_workbook(SRC, read_only=True)
    ph, phi, prev_rows = read_sheet(swb["前回"])
    ch, chi, curr_rows = read_sheet(swb["今回"])
    swb.close()

    problems = []   # チェックシートに出す異常
    infos = []

    # 必須ヘッダの存在チェック
    for name, hidx, sheet in ((ph, phi, "前回"), (ch, chi, "今回")):
        for h in KEY_HEADERS + ["実績番号", "測定日", "No", "線名", "線路名"]:
            if h not in hidx:
                problems.append(f"シート「{sheet}」にヘッダ「{h}」が見つかりません")
        for _, vh, fh in ITEMS:
            if vh not in hidx:
                problems.append(f"シート「{sheet}」にヘッダ「{vh}」が見つかりません")
            if fh and fh not in hidx:
                problems.append(f"シート「{sheet}」にヘッダ「{fh}」が見つかりません")
    if problems:
        raise SystemExit("ヘッダ不整合: " + "; ".join(problems))

    # 実績番号
    def jisseki(rows, hidx):
        vals = sorted({str(r[hidx["実績番号"]]).strip() for r in rows
                       if r[hidx["実績番号"]] not in (None, "")})
        return vals
    pj, cj = jisseki(prev_rows, phi), jisseki(curr_rows, chi)
    infos.append(("前回シート 実績番号", ", ".join(pj)))
    infos.append(("今回シート 実績番号", ", ".join(cj)))
    infos.append(("前回シート データ行数", len(prev_rows)))
    infos.append(("今回シート データ行数", len(curr_rows)))
    if len(pj) != 1:
        problems.append(f"前回シートに実績番号が{len(pj)}種類あります: {pj}")
    if len(cj) != 1:
        problems.append(f"今回シートに実績番号が{len(cj)}種類あります: {cj}")
    if pj and cj and pj[-1] >= cj[0]:
        problems.append(
            f"実績番号の新旧が逆の可能性: 前回={pj} / 今回={cj} (今回の方が新しい番号のはず)")

    # キー作成
    def make_key(row, hidx):
        parts = []
        for h in KEY_HEADERS:
            v = to_num(row[hidx[h]])
            parts.append("" if v is None else str(v))
        return "|".join(parts)

    prev_map = {}
    dup_prev = 0
    for r in prev_rows:
        k = make_key(r, phi)
        if k in prev_map:
            dup_prev += 1
        prev_map[k] = r
    seen_curr = set()
    dup_curr = 0
    for r in curr_rows:
        k = make_key(r, chi)
        if k in seen_curr:
            dup_curr += 1
        seen_curr.add(k)
    if dup_prev:
        problems.append(f"前回シートにキー重複が{dup_prev}件あります(線名・線別・キロ程等が同一の行)")
    if dup_curr:
        problems.append(f"今回シートにキー重複が{dup_curr}件あります(線名・線別・キロ程等が同一の行)")

    # 出力ワークブック
    wb = openpyxl.Workbook()
    ws_guide = wb.active
    ws_guide.title = "説明"
    ws_cmp = wb.create_sheet("比較")
    ws_ext = wb.create_sheet("抽出_要対応")
    ws_chk = wb.create_sheet("チェック")

    # ---- 比較シートヘッダ ----
    base_headers = ["No", "線名", "線路名", "キロ程", "判定", "該当項目",
                    "前回実績番号", "今回実績番号", "前回測定日", "今回測定日",
                    "WB", "本準側名"]
    headers = list(base_headers)
    item_col = {}  # 項目名 -> 先頭列(1始まり)
    for name, _, fh in ITEMS:
        item_col[name] = len(headers) + 1
        headers += [f"{name}前回", f"{name}今回", f"{name}差分"]
        if fh:
            headers.append(f"{name}目標F")
    ws_cmp.append(headers)

    unmatched_curr = []   # 今回にあって前回にないキー
    nonnum = {}           # 列ヘッダ -> 非数値セル数
    flag_counts = {n: 0 for n, _, fh in ITEMS if fh}
    band_low_counts = {n: 0 for n in flag_counts}
    band_high_counts = {n: 0 for n in flag_counts}

    fills = []  # (row, col, fill) を溜めて後で適用
    extract_src_rows = []  # 抽出対象の比較シート行番号

    out_r = 1
    for r in curr_rows:
        out_r += 1
        k = make_key(r, chi)
        p = prev_map.get(k)
        if p is None:
            unmatched_curr.append(to_num(r[chi["キロ程"]]))

        def cv(h):  # 今回の値
            return to_num(r[chi[h]])

        def pv(h):  # 前回の値
            return to_num(p[phi[h]]) if p is not None else None

        row_vals = [
            cv("No"), cv("線名"), cv("線路名"), cv("キロ程"),
            "", "",  # 判定・該当項目は後で
            pv("実績番号"), cv("実績番号"), pv("測定日"), cv("測定日"),
            cv("WB"), cv("本準側名"),
        ]
        hit_items = []
        worst = 0  # 0=なし 1=10-15 2=15以上
        for name, vh, fh in ITEMS:
            c0 = len(row_vals) + 1  # この項目の先頭列
            pval, cval = pv(vh), cv(vh)
            pf, cf = fnum(pval), fnum(cval)
            # 非数値検知(空欄以外で数値化できないもの)
            for tag, raw, f in (("前回", pval, pf), ("今回", cval, cf)):
                if raw not in (None, "") and f is None:
                    nonnum[f"{name}{tag}"] = nonnum.get(f"{name}{tag}", 0) + 1
            diff = round(cf - pf, 1) if (pf is not None and cf is not None) else None
            row_vals += [pval, cval, diff]
            flag = None
            if fh:
                flag = cv(fh)
                row_vals.append(flag)
            is_target = (fh is not None and fnum(flag) == 1.0)
            if is_target:
                flag_counts[name] += 1
                fills.append((out_r, c0 + 1, FILL_YELLOW))  # 今回値セル
                if diff is not None:
                    ad = abs(diff)
                    if ad >= TH_HIGH:
                        worst = max(worst, 2)
                        hit_items.append(name)
                        band_high_counts[name] += 1
                        fills.append((out_r, c0 + 1, FILL_RED))
                        fills.append((out_r, c0 + 2, FILL_RED))
                    elif ad >= TH_LOW:
                        worst = max(worst, 1)
                        hit_items.append(name)
                        band_low_counts[name] += 1
                        fills.append((out_r, c0 + 1, FILL_ORANGE))
                        fills.append((out_r, c0 + 2, FILL_ORANGE))
        if worst == 2:
            row_vals[4] = "15mm以上"
        elif worst == 1:
            row_vals[4] = "10mm以上15mm未満"
        row_vals[5] = "、".join(hit_items)
        if p is None:
            row_vals[5] = "(前回データなし)" if not hit_items else row_vals[5] + "、(前回データなし)"
        ws_cmp.append(row_vals)
        if worst == 2:
            fills.append((out_r, 5, FILL_RED))
            extract_src_rows.append(out_r)
        elif worst == 1:
            fills.append((out_r, 5, FILL_ORANGE))
            extract_src_rows.append(out_r)

    # 前回にあって今回にないキー
    curr_keys = {make_key(r, chi) for r in curr_rows}
    unmatched_prev = [to_num(pr[phi["キロ程"]]) for k, pr in prev_map.items()
                      if k not in curr_keys]

    # ---- 塗りつぶし適用 ----
    for rr, cc, f in fills:
        ws_cmp.cell(row=rr, column=cc).fill = f

    # ---- 書式(比較) ----
    ncols = len(headers)
    for c in range(1, ncols + 1):
        cell = ws_cmp.cell(row=1, column=c)
        cell.font = BOLD
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(vertical="center")
    # 数値表示形式
    fmt_map = {"キロ程": "0.0"}
    for name, _, fh in ITEMS:
        for suf in ("前回", "今回", "差分"):
            fmt_map[f"{name}{suf}"] = "0.0"
    col_of = {h: i + 1 for i, h in enumerate(headers)}
    maxr = ws_cmp.max_row
    for h, fmt in fmt_map.items():
        c = col_of[h]
        for rr in range(2, maxr + 1):
            ws_cmp.cell(row=rr, column=c).number_format = fmt
    # 列幅
    widths = {"No": 6, "線名": 11, "線路名": 9, "キロ程": 10, "判定": 16,
              "該当項目": 18, "前回実績番号": 12, "今回実績番号": 12,
              "前回測定日": 11, "今回測定日": 11, "WB": 5, "本準側名": 9}
    for h, w in widths.items():
        ws_cmp.column_dimensions[get_column_letter(col_of[h])].width = w
    for h in col_of:
        if h not in widths:
            ws_cmp.column_dimensions[get_column_letter(col_of[h])].width = 11
    ws_cmp.auto_filter.ref = f"A1:{get_column_letter(ncols)}{maxr}"
    ws_cmp.freeze_panes = "G2"

    # ---- 抽出_要対応 ----
    ws_ext.append(headers)
    for c in range(1, ncols + 1):
        cell = ws_ext.cell(row=1, column=c)
        cell.font = BOLD
        cell.fill = FILL_HEADER
    fill_lookup = {}
    for rr, cc, f in fills:
        fill_lookup[(rr, cc)] = f  # 後勝ち(RED/ORANGEがYELLOWを上書き)
    er = 1
    for src in extract_src_rows:
        er += 1
        for c in range(1, ncols + 1):
            scell = ws_cmp.cell(row=src, column=c)
            dcell = ws_ext.cell(row=er, column=c, value=scell.value)
            dcell.number_format = scell.number_format
            f = fill_lookup.get((src, c))
            if f is not None:
                dcell.fill = f
    for h, w in widths.items():
        ws_ext.column_dimensions[get_column_letter(col_of[h])].width = w
    for h in col_of:
        if h not in widths:
            ws_ext.column_dimensions[get_column_letter(col_of[h])].width = 11
    if er > 1:
        ws_ext.auto_filter.ref = f"A1:{get_column_letter(ncols)}{er}"
    ws_ext.freeze_panes = "G2"

    # ---- チェックシート ----
    ws_chk.column_dimensions["A"].width = 42
    ws_chk.column_dimensions["B"].width = 60
    r = 1
    def put(label, value, warn=False, ok=False):
        nonlocal r
        a = ws_chk.cell(row=r, column=1, value=label)
        b = ws_chk.cell(row=r, column=2, value=value)
        if warn:
            b.fill = FILL_WARN
            b.font = BOLD
        elif ok:
            b.fill = FILL_OK
        r += 1

    ws_chk.cell(row=r, column=1, value="■ 作成結果チェック").font = BOLD
    r += 2
    put("作成日時", datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S"))
    for label, v in infos:
        put(label, v)
    put("比較シート 行数(今回と同数のはず)", maxr - 1,
        warn=(maxr - 1 != len(curr_rows)), ok=(maxr - 1 == len(curr_rows)))
    put("今回にあって前回にない箇所(前回データなし)", len(unmatched_curr),
        warn=len(unmatched_curr) > 0, ok=len(unmatched_curr) == 0)
    if unmatched_curr:
        put("  └ 該当キロ程(先頭50件)", ", ".join(str(x) for x in unmatched_curr[:50]))
    put("前回にあって今回にない箇所", len(unmatched_prev),
        warn=len(unmatched_prev) > 0, ok=len(unmatched_prev) == 0)
    if unmatched_prev:
        put("  └ 該当キロ程(先頭50件)", ", ".join(str(x) for x in unmatched_prev[:50]))
    total_nonnum = sum(nonnum.values())
    put("数値にできなかった測定値セル", total_nonnum,
        warn=total_nonnum > 0, ok=total_nonnum == 0)
    for h, n in sorted(nonnum.items()):
        put(f"  └ {h}", n, warn=True)
    r += 1
    ws_chk.cell(row=r, column=1, value="■ 検出件数").font = BOLD
    r += 1
    put("今回 目標値F=1 の箇所(項目別)",
        "、".join(f"{k}:{v}件" for k, v in flag_counts.items() if v) or "なし")
    low_total = sum(band_low_counts.values())
    high_total = sum(band_high_counts.values())
    put("うち 変位進み量 10mm以上15mm未満",
        ("、".join(f"{k}:{v}件" for k, v in band_low_counts.items() if v) or "0件"))
    put("うち 変位進み量 15mm以上",
        ("、".join(f"{k}:{v}件" for k, v in band_high_counts.items() if v) or "0件"))
    put("抽出_要対応シートの行数", len(extract_src_rows))
    r += 1
    ws_chk.cell(row=r, column=1, value="■ 異常・警告").font = BOLD
    r += 1
    if problems:
        for pmsg in problems:
            put("警告", pmsg, warn=True)
    else:
        put("警告", "なし(正常に作成されました)", ok=True)

    # ---- 説明シート ----
    ws_guide.column_dimensions["A"].width = 100
    guide = [
        ("■ このブックについて", True),
        ("四半期ごとの軌道検査データ(前回・今回)を比較し、目標値F超過箇所と変位進み量を色付けで示します。", False),
        ("", False),
        ("■ シート構成", True),
        ("・比較 …… 今回データの各キロ程に前回値・今回値・差分(今回−前回)を横1行で並べたもの。", False),
        ("・抽出_要対応 …… 変位進み量が10mm以上の行だけを抜き出したもの。", False),
        ("・チェック …… 作成時の検証結果。警告が赤で表示されたらデータを確認してください。", False),
        ("", False),
        ("■ 色の意味(比較・抽出_要対応シート)", True),
        ("・黄色 …… 今回の目標値F(目標F)が「1」の測定値セル。", False),
        ("・オレンジ …… 目標値F=1 かつ 変位進み量が10mm以上15mm未満 → 30日以内に調査、変位進み確認時は60日以内に処置。", False),
        ("・赤 …… 目標値F=1 かつ 変位進み量が15mm以上 → 15日以内に調査、変位進み確認時は30日以内に処置。", False),
        ("・「判定」列に該当区分、「該当項目」列にどの測定項目で該当したかが入ります。", False),
        ("", False),
        ("■ 変位進み量の定義", True),
        ("・変位進み量 = |今回値 − 前回値| (差分の絶対値)。差分列には符号付きの値(今回−前回)を表示しています。", False),
        ("・前回と今回はキロ程(+線名コード・線別コード・番線番号・本準側コード)で突き合わせています。", False),
        ("", False),
        ("■ フィルタの使い方", True),
        ("・比較シートの「判定」列でフィルタすると、10mm以上15mm未満/15mm以上の行だけを表示できます。", False),
        ("・すべて数値として格納しているので、キロ程や差分での並べ替え・数値フィルタも使えます。", False),
        ("", False),
        ("■ 次回以降(データ差し替え時)の作り方", True),
        ("・同梱のVBAマクロ「比較シート作成.bas」を自分のブックに取り込み、前回・今回シートを差し替えてマクロを実行すると、", False),
        ("  比較・抽出_要対応・チェックの3シートが自動で作り直されます。手順は同梱の「README(手順書).txt」参照。", False),
        ("", False),
        ("■ 軽量化のポイント", True),
        ("・数値を文字列ではなく数値として保存し、数式や条件付き書式を使わず値と色だけにしてあるため、動作が軽くなります。", False),
        ("・元データの不要な列・古い比較シートを残さないことでも容量を減らせます。", False),
    ]
    for i, (text, bold) in enumerate(guide, 1):
        c = ws_guide.cell(row=i, column=1, value=text)
        if bold:
            c.font = BOLD

    # ---- 前回/今回シート(数値化して収録) ----
    swb = openpyxl.load_workbook(SRC, read_only=True)
    meas_fmt_headers = set()
    for name, vh, fh in ITEMS:
        meas_fmt_headers.add(vh)
    meas_fmt_headers.add("キロ程")
    for sname in ("前回", "今回"):
        sws = swb[sname]
        dws = wb.create_sheet(sname)
        rows = sws.iter_rows(values_only=True)
        hdr = next(rows)
        hdr = [("" if h is None else str(h).strip()) for h in hdr]
        dws.append(hdr)
        fmt_cols = {i + 1 for i, h in enumerate(hdr) if h in meas_fmt_headers}
        for c in range(1, len(hdr) + 1):
            dws.cell(row=1, column=c).font = BOLD
            dws.cell(row=1, column=c).fill = FILL_HEADER
        rr = 1
        for row in rows:
            if not any(c not in (None, "") for c in row):
                continue
            rr += 1
            dws.append([to_num(v) for v in row])
            for c in fmt_cols:
                dws.cell(row=rr, column=c).number_format = "0.0"
        dws.auto_filter.ref = f"A1:{get_column_letter(len(hdr))}{rr}"
        dws.freeze_panes = "A2"
    swb.close()

    print("saving ...")
    wb.save(OUT)
    print("done:", OUT)
    print("problems:", problems)
    print("flag:", flag_counts, "low:", band_low_counts, "high:", band_high_counts)
    print("extract rows:", len(extract_src_rows))


if __name__ == "__main__":
    build()
