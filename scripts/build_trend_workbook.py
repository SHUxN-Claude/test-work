#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""年度推移管理ツール(動揺データ)のExcelブックを生成する。

4月〜翌3月を1年度として毎月の検査PDFを取り込み、
同一箇所(キロ程が設定距離以内)ごとの月別推移を管理するブック。

  設定       … 目標値/基準値/黄色強調値/同一箇所とみなす距離/年度
  PDF貼付    … 取込ボタン(VBA)
  推移表     … 箇所ごとの月別推移(4月〜3月) + 施工記録・前年度値
  グラフ     … 箇所選択式の折れ線 + ワースト5
  全データ   … 取り込んだ全行(月列付き)
  ①基準値超過/②目標値超過 … 月をまたいだ抽出一覧
  施工記録   … 線別+キロ程で施工を記録(手入力)
  取込履歴   … 月別の取込状況
  前年度実績 … 新年度開始時に前年度の推移表を退避

使い方:
  python build_trend_workbook.py -o テンプレート.xlsx          # 空のテンプレート
  python build_trend_workbook.py --demo 実PDF.pdf -o デモ.xlsx # 疑似データ入りデモ
"""
import argparse
import random

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from pdf_to_excel import parse_pdf

FONT = "游ゴシック"
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", start_color="D9E1F2")
YELLOW_FILL = PatternFill("solid", start_color="FFFF00")
RED_FILL = PatternFill("solid", start_color="FFC7CE")
RED_FONT = Font(name=FONT, color="9C0006", bold=True)
INPUT_FILL = PatternFill("solid", start_color="FFF2CC")

DATA_START = 5
CF_LAST = 6000
MONTH_LABELS = ["4月", "5月", "6月", "7月", "8月", "9月",
                "10月", "11月", "12月", "1月", "2月", "3月"]

DATA_HEADERS = ["NO", "列車速度\n(km/h)", "時刻\n(H:M:S)", "キロ程\n(km)",
                "上下動\n(G)", "左右動\n(G)", "超過印", "備考",
                "測定日", "線別", "保線技術センター", "元ファイル", "月"]
DATA_WIDTHS = [6, 10, 10, 10, 9, 9, 7, 18, 12, 8, 16, 24, 7]

TREND_HEADERS = (["箇所No", "線別", "区分", "代表キロ程\n(km)",
                  "範囲(自)", "範囲(至)", "備考\n(PDF自動)", "メモ\n(手入力)"]
                 + MONTH_LABELS
                 + ["年間最大", "前年度値", "施工日", "施工内容"])
TREND_WIDTHS = [7, 7, 8, 11, 9, 9, 14, 18] + [7] * 12 + [8, 8, 11, 24]
COL_MEMO = 8                  # H列 = メモ(手入力)
TREND_MONTH_COL0 = 9          # I列 = 4月


def fiscal_index(month):
    """暦月(1-12) → 年度内の月番号(4月=1 … 3月=12)"""
    return (month - 4) % 12 + 1


def header_row(ws, headers, widths, row=4):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, bold=True, size=10)
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def add_value_cf(ws, rng, anchor):
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f"AND(ISNUMBER({anchor}),{anchor}>=基準値)"],
        fill=RED_FILL, font=RED_FONT, stopIfTrue=True))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f"AND(ISNUMBER({anchor}),{anchor}>=黄色強調値,{anchor}<基準値)"],
        fill=YELLOW_FILL, stopIfTrue=True))


def note(ws, row, col, text, **kw):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name=FONT, size=kw.pop("size", 10), **kw)
    return c


# ---------------------------------------------------------------- 設定
def build_settings(ws):
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 72
    note(ws, 1, 2, "■ 設定シート(職場ごとにここだけ変更してください)",
         bold=True, size=12)
    items = [
        ("目標値", 0.20, "この値以上を「②目標値超過」に抽出します。"),
        ("基準値", 0.25, "この値以上を「①基準値超過」に抽出します。"),
        ("黄色強調値", 0.24, "この値以上のセルを黄色で強調します。"),
        ("同一箇所とみなす距離(m)", 20,
         "キロ程がこの距離以内で連続する検出点を同じ箇所として推移をまとめます。"),
        ("年度(開始年)", 2026, "例: 2026 → 2026年4月〜2027年3月を1年度として扱います。"),
    ]
    for i, (name, val, desc) in enumerate(items):
        r = 3 + i
        note(ws, r, 2, name, bold=True)
        c = ws.cell(row=r, column=3, value=val)
        c.font = Font(name=FONT, color="0000FF", bold=True)
        c.fill = INPUT_FILL
        c.border = BORDER
        c.number_format = "0.00" if i < 3 else "0"
        note(ws, r, 4, desc)
    tips = [
        "",
        "【凡例】 色付き(うすい橙)のセルが入力欄です。数値のみ変更してください。",
        "",
        "【例】 目標値0.25/基準値0.30の職場 → 0.25 / 0.30 / 0.29 に変更",
        "",
        "※ 値を変更したら「PDF貼付」シートの「再計算」ボタンを押してください。",
        "※ 「同一箇所とみなす距離」を変えると、4月以降の全データを新しい距離で",
        "   グループ分けし直します(過去の月も含めて再計算されます)。",
        "   距離を大きくすると、それまで別々だった箇所が1つにまとまることがあります。",
        "   施工記録は「施工記録」シートにキロ程で記録するため、組み替え後も自動で追従します。",
    ]
    for i, t in enumerate(tips):
        note(ws, 9 + i, 2, t)


# ---------------------------------------------------------------- PDF貼付
def build_paste(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 14
    note(ws, 2, 2, "■ PDF貼付シート", bold=True, size=14)
    warn = [
        "【重要】いまのファイル形式(.xlsx)ではボタンは使えません。",
        "　　　　下の「初回セットアップ」を先に行ってください。",
        "　　　　(Alt+F8 のマクロ一覧が空なのは、セットアップがまだのためです)",
    ]
    for i, s in enumerate(warn):
        note(ws, 4 + i, 2, s, color="C00000", bold=(i == 0))
    setup = [
        "【初回セットアップ(最初に1回だけ・2〜3分)】",
        "  1. Alt+F11 を押してVBA画面を開く",
        "  2. メニューの「ファイル」→「ファイルのインポート」で、配布された vba フォルダの",
        "     modTrend.bas / modEngine.bas / modEmbedded.bas を1つずつ、計3回インポートする",
        "     (単月ツールの modMain.bas は入れないでください。マクロ名が重複します)",
        "  3. Alt+F11 でExcelに戻り、F12(名前を付けて保存)を押す",
        "  4. ファイルの種類を「Excel マクロ有効ブック (*.xlsm)」に変えて保存する ← 重要",
        "  5. Alt+F8 →「初期設定」を実行すると、このシートにボタンが設置されます",
    ]
    for i, s in enumerate(setup):
        note(ws, 9 + i, 2, s, bold=s.startswith("【"))
    note(ws, 19, 2, "▼ セットアップ後、ここにボタンが表示されます ▼", color="808080")
    steps = [
        "【毎月の使い方】",
        "  1. 「PDFファイルを選択して取込」ボタンで、その月のPDFを選ぶ",
        "     (またはこのシートの下にPDFを貼り付けて保存 →「貼り付けたPDFを取り込む」)",
        "  2. 「全データ」に追加され、「推移表」「グラフ」「①②」が自動で更新されます",
        "  3. 同じ月のPDFを取り込み直すと、その月の分だけ入れ替わります(二重計上しません)",
        "",
        "【設定を変えたとき】 「再計算」ボタンを押してください(PDFは読み直しません)",
        "【施工したとき】 「施工記録」シートに線別・キロ程・施工日・内容を入力し「再計算」",
        "【新年度になったら】 このファイルをコピーし、コピー側で Alt+F8 →「新年度開始」を実行",
        "",
        "※ .xlsm で保存し直さずに閉じると、インポートしたマクロは消えてしまいます。",
        "※ OneDrive上に保存していると「貼り付けたPDFの取込」が使えない場合があります。",
        "   その場合はローカル(デスクトップ等)に保存するか、ファイル選択取込をご利用ください。",
    ]
    for i, s in enumerate(steps):
        note(ws, 23 + i, 2, s, bold=s.startswith("【"))
    note(ws, 40, 2, "▼ この下にPDFを貼り付けてください ▼",
         bold=True, size=11, color="1F4E79")


# ---------------------------------------------------------------- データ系
def build_data_sheet(ws, note_formula):
    header_row(ws, DATA_HEADERS, DATA_WIDTHS)
    a2 = ws.cell(row=2, column=1, value=note_formula)
    a2.font = Font(name=FONT, size=10, color="1F4E79")
    note(ws, 2, 6, "該当件数：")
    c = ws.cell(row=2, column=7, value=f"=COUNT(E{DATA_START}:F{CF_LAST})")
    c.font = Font(name=FONT, size=10, bold=True)
    add_value_cf(ws, f"E{DATA_START}:F{CF_LAST}", f"E{DATA_START}")
    fmts = {"B": "0.0", "C": "@", "D": "0.000", "E": "0.00", "F": "0.00"}
    for col, fmt in fmts.items():
        for r in range(DATA_START, CF_LAST + 1):
            ws[f"{col}{r}"].number_format = fmt


def build_trend_sheet(ws):
    header_row(ws, TREND_HEADERS, TREND_WIDTHS)
    note(ws, 2, 1, "箇所ごとの月別推移(値はその箇所・その月の最大値)。"
         "「−」は取込済みの月にその箇所の検出がなかった(=0.15G未満に収まっていた)ことを表します。",
         color="1F4E79")
    note(ws, 3, 1, "※ 箇所Noは再計算のたびにキロ程順で振り直します。"
         "「メモ」列は直接入力できます(再計算しても消えません)。"
         "施工の記録は「施工記録」シートへ。",
         color="808080", size=9)
    first = get_column_letter(TREND_MONTH_COL0) + str(DATA_START)          # I5
    last = get_column_letter(TREND_MONTH_COL0 + 13) + str(CF_LAST)         # V6000
    add_value_cf(ws, f"{first}:{last}", first)
    for col in ("D", "E", "F"):
        for r in range(DATA_START, CF_LAST + 1):
            ws[f"{col}{r}"].number_format = "0.000"
    # メモ列は手入力欄と分かるよう色を付けておく
    memo_col = get_column_letter(COL_MEMO)
    for r in range(DATA_START, CF_LAST + 1):
        ws[f"{memo_col}{r}"].fill = INPUT_FILL


def build_repair_sheet(ws):
    ws.column_dimensions["A"].width = 3
    headers = ["線別", "キロ程(km)", "施工日", "施工内容"]
    widths = [8, 12, 12, 40]
    for i, (h, w) in enumerate(zip(headers, widths), start=2):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name=FONT, bold=True, size=10)
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    note(ws, 1, 2, "■ 施工記録(手入力)", bold=True, size=12)
    note(ws, 2, 2, "施工したら1行ずつ入力してください。キロ程は施工箇所の代表位置でかまいません"
         "(「同一箇所とみなす距離」の範囲で推移表の箇所に自動でひも付きます)。入力後「再計算」。")
    note(ws, 3, 2, "入力例: 下線 / 105.390 / 2026/06/15 / むら直し施工 "
         "(この行は例です。実際の入力は5行目から)", color="808080", size=9)
    for r in range(5, 205):
        for col in range(2, 6):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).fill = INPUT_FILL
        ws.cell(row=r, column=3).number_format = "0.000"
        ws.cell(row=r, column=4).number_format = "yyyy/mm/dd"


def build_memo_sheet(ws):
    """手入力メモの保管シート。

    推移表の「メモ」列に直接入力された内容は、再計算の直前にこのシートへ
    退避され(線別+キロ程で記録)、再計算後に位置の突き合わせで貼り直される。
    このシートに直接書いてもよい。
    """
    ws.column_dimensions["A"].width = 3
    headers = ["線別", "キロ程(km)", "メモ"]
    widths = [8, 12, 46]
    for i, (h, w) in enumerate(zip(headers, widths), start=2):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name=FONT, bold=True, size=10)
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    note(ws, 1, 2, "■ 箇所メモ(手入力の保管場所)", bold=True, size=12)
    note(ws, 2, 2, "推移表の「メモ」列に入力した内容は、再計算のたびにここへ自動保存され、"
         "計算後に同じ位置へ貼り直されます。ここに直接書いても構いません。")
    note(ws, 3, 2, "入力例: 下線 / 74.212 / 巌Ｔ 出口の継目部 "
         "(この行は例です。実際の入力は5行目から)", color="808080", size=9)
    for r in range(5, 405):
        for col in range(2, 5):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).fill = INPUT_FILL
        ws.cell(row=r, column=3).number_format = "0.000"


def build_history_sheet(ws):
    ws.column_dimensions["A"].width = 3
    headers = ["月", "線別", "測定日", "元ファイル", "行数", "取込日時"]
    widths = [7, 8, 12, 30, 8, 18]
    for i, (h, w) in enumerate(zip(headers, widths), start=2):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name=FONT, bold=True, size=10)
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    note(ws, 1, 2, "■ 取込履歴(自動で記録されます)", bold=True, size=12)
    note(ws, 2, 2, "取込済みの月にPDFへ載っていない箇所は「0.15G未満に収まっていた」と判断し、"
         "推移表に「−」を表示します。")


def build_prev_sheet(ws):
    header_row(ws, TREND_HEADERS, TREND_WIDTHS)
    note(ws, 1, 1, "■ 前年度実績(「新年度開始」を実行すると前年度の推移表がここに退避されます)",
         bold=True, size=12)
    note(ws, 2, 1, "推移表の「前年度値」列は、このシートの年間最大をキロ程の突き合わせで表示します。",
         color="1F4E79")


CHART_SEL_ROW = 41      # 選択箇所の系列(41=実測 / 42=目標値 / 43=基準値)
CHART_WORST_ROW = 51    # ワースト5の系列(51〜55) / 56=基準値


def build_chart_sheet(ws):
    """グラフシートの骨組みと2つの折れ線グラフを作る。

    グラフは固定の行範囲(40〜43行 / 50〜56行)を参照するため、
    VBA側はその範囲に値を書き込むだけでグラフが切り替わる。
    """
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    note(ws, 1, 2, "■ 推移グラフ", bold=True, size=14)
    note(ws, 2, 2, "箇所No →")
    sel = ws.cell(row=2, column=3)
    sel.fill = INPUT_FILL
    sel.border = BORDER
    sel.font = Font(name=FONT, bold=True, color="0000FF")
    note(ws, 2, 4, "← 箇所Noを選び「再計算」ボタンを押すと、上のグラフが切り替わります。",
         size=9, color="808080")
    note(ws, 38, 2, "▼ グラフ用データ(自動生成・編集不要)", size=9, color="808080")

    # --- 系列ラベルと月見出し(値はVBA/デモが書き込む) ---
    ws.cell(row=40, column=1, value="選択箇所")
    ws.cell(row=50, column=1, value="ワースト5(最新月)")
    for j, label in enumerate(MONTH_LABELS):
        ws.cell(row=40, column=2 + j, value=label)
        ws.cell(row=50, column=2 + j, value=label)
    ws.cell(row=41, column=1, value="(箇所を選んでください)")
    ws.cell(row=42, column=1, value="目標値")
    ws.cell(row=43, column=1, value="基準値")
    for k in range(5):
        ws.cell(row=51 + k, column=1, value=f"(ワースト{k + 1})")
    ws.cell(row=56, column=1, value="基準値")

    ch = LineChart()
    ch.title = "選択した箇所の推移"
    ch.y_axis.title = "全振幅(G)"
    ch.height, ch.width = 9, 17
    ch.add_data(Reference(ws, min_col=1, max_col=13, min_row=41, max_row=43),
                from_rows=True, titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=2, max_col=13, min_row=40))
    ws.add_chart(ch, "B4")

    ch2 = LineChart()
    ch2.title = "ワースト5(最新月)の推移"
    ch2.y_axis.title = "全振幅(G)"
    ch2.height, ch2.width = 9, 17
    ch2.add_data(Reference(ws, min_col=1, max_col=13, min_row=51, max_row=56),
                 from_rows=True, titles_from_data=True)
    ch2.set_categories(Reference(ws, min_col=2, max_col=13, min_row=50))
    ws.add_chart(ch2, "B24")

    dv = DataValidation(type="list", formula1=f"=推移表!$A$5:$A${CF_LAST}",
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(sel)


# ---------------------------------------------------------------- 本体
def build_book():
    wb = Workbook()
    ws = wb.active
    ws.title = "設定"
    build_settings(ws)
    for name, cell in (("目標値", "$C$3"), ("基準値", "$C$4"),
                       ("黄色強調値", "$C$5")):
        wb.defined_names.add(DefinedName(name, attr_text=f"設定!{cell}"))
    build_paste(wb.create_sheet("PDF貼付"))
    build_trend_sheet(wb.create_sheet("推移表"))
    build_chart_sheet(wb.create_sheet("グラフ"))
    build_data_sheet(wb.create_sheet("全データ"), '="取り込んだ全データ(全月分)"')
    build_data_sheet(
        wb.create_sheet("①基準値超過"),
        '="抽出条件： 上下動もしくは左右動が 基準値("&TEXT(基準値,"0.00")&"G)以上(全月分)"')
    build_data_sheet(
        wb.create_sheet("②目標値超過"),
        '="抽出条件： 目標値("&TEXT(目標値,"0.00")&"G)以上〜基準値("&TEXT(基準値,"0.00")'
        '&"G)未満(全月分)　※ "&TEXT(黄色強調値,"0.00")&"G以上は黄色"')
    build_memo_sheet(wb.create_sheet("箇所メモ"))
    build_repair_sheet(wb.create_sheet("施工記録"))
    build_history_sheet(wb.create_sheet("取込履歴"))
    build_prev_sheet(wb.create_sheet("前年度実績"))
    return wb


# ---------------------------------------------------------------- グループ化
def group_points(points, tol_km):
    """points: (line, kind, km, val, fiscal_idx, biko) のリスト。
    線別・区分ごとにキロ程順で並べ、隣との差が tol_km 以内なら同一箇所に
    つなげる(鎖方式)。戻り値: クラスタのリスト。"""
    clusters = []
    for key in sorted({(p[0], p[1]) for p in points}):
        pts = sorted((p for p in points if (p[0], p[1]) == key),
                     key=lambda p: p[2])
        cur = None
        for p in pts:
            if cur is None or p[2] - cur["kms"][-1] > tol_km:
                cur = {"line": p[0], "kind": p[1], "kms": [], "months": {},
                       "biko": "", "rep": None, "maxval": -1}
                clusters.append(cur)
            cur["kms"].append(p[2])
            mi = p[4]
            cur["months"][mi] = max(cur["months"].get(mi, 0), p[3])
            if p[3] > cur["maxval"]:
                cur["maxval"] = p[3]
                cur["rep"] = p[2]
            if p[5] and not cur["biko"]:
                cur["biko"] = p[5]
    clusters.sort(key=lambda c: (c["line"], c["rep"], c["kind"]))
    return clusters


# ---------------------------------------------------------------- 書き出し
def match_at(items, line, lo, hi, tol):
    """線別が一致し、キロ程が [lo-tol, hi+tol] に入る項目を拾う。"""
    return [it for it in items
            if it[0] == line and lo - tol <= it[1] <= hi + tol]


def write_memo_sheet(wb, memos):
    """箇所メモシートへ書き出す(線別・キロ程・メモ)。"""
    ws = wb["箇所メモ"]
    for i, (line, km, text) in enumerate(memos):
        r = 5 + i
        for col, v in enumerate((line, km, text), start=2):
            c = ws.cell(row=r, column=col, value=v)
            c.font = Font(name=FONT, size=10)


def write_trend(wb, clusters, imported, tol, repairs, memos=()):
    """推移表を書き出す。imported に含まれる月で検出がない箇所は「−」。

    memos: (線別, キロ程, メモ) の一覧。位置が箇所の範囲に入るものを
    「メモ(手入力)」列へ貼り直す。グループの併合で複数該当した場合は
    どれも失われないよう「 / 」でつなぐ。
    """
    ws = wb["推移表"]
    month_end = get_column_letter(TREND_MONTH_COL0 + 11)
    month_start = get_column_letter(TREND_MONTH_COL0)
    for i, cl in enumerate(clusters):
        r = DATA_START + i
        lo, hi = cl["kms"][0], cl["kms"][-1]
        rep_d, rep_txt = None, ""
        for line, km, d, txt in repairs:
            if line == cl["line"] and lo - tol <= km <= hi + tol:
                rep_d, rep_txt = d, txt
        hit = match_at(memos, cl["line"], lo, hi, tol)
        memo = " / ".join(dict.fromkeys(m[2] for m in hit if m[2]))
        vals = [i + 1, cl["line"], cl["kind"], cl["rep"], lo, hi,
                cl["biko"], memo]
        vals += [cl["months"].get(mi, "−" if mi in imported else None)
                 for mi in range(1, 13)]
        vals += [f"=MAX({month_start}{r}:{month_end}{r})", None,
                 rep_d, rep_txt]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
            if col in (1, 2, 3) or TREND_MONTH_COL0 <= col < TREND_MONTH_COL0 + 12:
                c.alignment = Alignment(horizontal="center")
            if TREND_MONTH_COL0 <= col <= TREND_MONTH_COL0 + 13:
                c.number_format = "0.00"
            if col == COL_MEMO:
                c.fill = INPUT_FILL


def write_chart_data(wb, clusters, imported, target, limit, selected=None):
    """グラフが参照する固定行(40〜43 / 50〜56)に値を書き込む。"""
    ws = wb["グラフ"]
    if not clusters or not imported:
        return
    if selected is None or not (1 <= selected <= len(clusters)):
        selected = 1
    ws.cell(row=2, column=3, value=selected)

    cl = clusters[selected - 1]
    ws.cell(row=CHART_SEL_ROW, column=1,
            value=f"No{selected} {cl['line']} {cl['rep']:.3f}({cl['kind']})")
    for j in range(12):
        ws.cell(row=CHART_SEL_ROW, column=2 + j, value=cl["months"].get(j + 1))
        ws.cell(row=CHART_SEL_ROW + 1, column=2 + j, value=target)
        ws.cell(row=CHART_SEL_ROW + 2, column=2 + j, value=limit)

    latest = imported[-1]
    top = sorted([c for c in clusters if latest in c["months"]],
                 key=lambda c: c["months"][latest], reverse=True)[:5]
    for k in range(5):
        r = CHART_WORST_ROW + k
        if k < len(top):
            c = top[k]
            no = clusters.index(c) + 1
            ws.cell(row=r, column=1,
                    value=f"No{no} {c['rep']:.3f}({c['kind']})")
            for j in range(12):
                ws.cell(row=r, column=2 + j, value=c["months"].get(j + 1))
        else:
            ws.cell(row=r, column=1, value=f"(ワースト{k + 1})")
    for j in range(12):
        ws.cell(row=CHART_WORST_ROW + 5, column=2 + j, value=limit)


# ---------------------------------------------------------------- デモ
def synth_month(base_rows, keep, jitter, scale, seed, exclude=(), extra=()):
    rnd = random.Random(seed)
    out = []
    for r in base_rows:
        if any(lo <= r["km"] <= hi for lo, hi in exclude):
            continue
        if rnd.random() > keep:
            continue
        v = r["ud"] if r["ud"] is not None else r["sa"]
        v = round(v * scale * rnd.uniform(0.9, 1.08), 2)
        if v < 0.15:
            continue
        row = dict(r)
        row["km"] = round(r["km"] + rnd.uniform(-jitter, jitter), 3)
        if r["ud"] is not None:
            row["ud"], row["sa"] = v, None
        else:
            row["ud"], row["sa"] = None, v
        row["ud_star"] = row["ud"] is not None and v >= 0.25
        row["sa_star"] = row["sa"] is not None and v >= 0.25
        out.append(row)
    for km, kind, v, biko in extra:
        out.append(dict(no=0, speed=80.0, time="", km=km,
                        ud=v if kind == "ud" else None,
                        ud_star=(kind == "ud" and v >= 0.25),
                        sa=v if kind == "sa" else None,
                        sa_star=(kind == "sa" and v >= 0.25),
                        biko=biko))
    out.sort(key=lambda r: r["km"])
    for i, r in enumerate(out, 1):
        r["no"] = i
    return out


def write_data_rows(ws, entries):
    r = DATA_START
    for month_label, meta, row in entries:
        vals = [row["no"], row["speed"], row["time"], row["km"],
                row["ud"], row["sa"],
                "*" if (row["ud_star"] or row["sa_star"]) else "",
                row["biko"], meta["date"], meta["line"], meta["center"],
                meta["file"], month_label]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
            if i in (1, 3, 7, 9, 10, 13):
                c.alignment = Alignment(horizontal="center")
        r += 1


def fill_demo(wb, pdf_path):
    meta, june = parse_pdf(pdf_path)
    story = [(105.380, 105.400)]                      # 悪化→施工→改善の物語箇所
    months = {
        3: [dict(r) for r in june],                                     # 6月(実データ)
        1: synth_month(june, 0.55, 0.008, 0.93, 41, exclude=story,
                       extra=[(105.390, "ud", 0.20, "")]),              # 4月
        2: synth_month(june, 0.65, 0.008, 0.97, 42, exclude=story,
                       extra=[(105.393, "ud", 0.22, "")]),              # 5月
        4: synth_month(june, 0.70, 0.008, 0.90, 43, exclude=story,
                       extra=[(105.388, "ud", 0.16, "")]),              # 7月
    }
    metas = {
        1: dict(meta, date="2026/04/03", file="(デモ疑似データ)4月.pdf"),
        2: dict(meta, date="2026/05/08", file="(デモ疑似データ)5月.pdf"),
        3: dict(meta, date="2026/06/04",
                file="(実データ)202603_2026.06.04.pdf"),
        4: dict(meta, date="2026/07/02", file="(デモ疑似データ)7月.pdf"),
    }
    tol = 0.020
    tgt, lim = 0.20, 0.25

    all_entries, over, tgt_rows, points = [], [], [], []
    for mi in sorted(months):
        label = MONTH_LABELS[mi - 1]
        for row in months[mi]:
            e = (label, metas[mi], row)
            all_entries.append(e)
            v = row["ud"] if row["ud"] is not None else row["sa"]
            if v is None:
                continue
            if v >= lim:
                over.append(e)
            elif v >= tgt:
                tgt_rows.append(e)
            kind = "上下動" if row["ud"] is not None else "左右動"
            points.append((metas[mi]["line"], kind, row["km"], v, mi,
                           row["biko"]))

    write_data_rows(wb["全データ"], all_entries)
    write_data_rows(wb["①基準値超過"], over)
    write_data_rows(wb["②目標値超過"], tgt_rows)

    # 施工記録(デモ)
    ws = wb["施工記録"]
    for col, v in zip(range(2, 6), ("下線", 105.390, "2026/06/15", "むら直し施工")):
        ws.cell(row=5, column=col, value=v)

    # 取込履歴
    ws = wb["取込履歴"]
    for i, mi in enumerate(sorted(months)):
        vals = [MONTH_LABELS[mi - 1], metas[mi]["line"], metas[mi]["date"],
                metas[mi]["file"], len(months[mi]), "(デモ)"]
        for col, v in enumerate(vals, start=2):
            c = ws.cell(row=5 + i, column=col, value=v)
            c.border = BORDER
            c.font = Font(name=FONT, size=10)

    clusters = group_points(points, tol)
    imported = sorted(months)
    repairs = [("下線", 105.390, "2026/06/15", "むら直し施工")]
    # 手入力メモの例(PDFに備考がない箇所へ後から書き足したという想定)
    memos = [("下線", 105.390, "継目部 要注意"),
             ("下線", 120.400, "踏切前後")]
    write_memo_sheet(wb, memos)
    write_trend(wb, clusters, imported, tol, repairs, memos)

    story_no = next(i + 1 for i, c in enumerate(clusters)
                    if abs(c["rep"] - 105.385) < 0.02 and c["kind"] == "上下動")
    story = clusters[story_no - 1]
    write_chart_data(wb, clusters, imported, tgt, lim, selected=story_no)
    print(f"デモ: {len(clusters)}箇所 / 物語の箇所No={story_no} "
          f"(4月0.20→5月0.22→6月{story['months'].get(3)}→7月{story['months'].get(4)})")


def read_carry(path, tol_m=20):
    """既存ブックから手入力分(メモ・施工記録)を読み出す。

    メモは「箇所メモ」シートを正本とし、推移表の「メモ」列からは
    **まだ箇所メモに無いもの(=推移表に直接入力された新しいメモ)だけ**を拾う。
    自動表示されたメモをそのまま拾い直すと、再計算のたびに記録位置が
    増えていき(メモの増殖)、やがて無関係な箇所まで巻き込むため。
    """
    from openpyxl import load_workbook
    tol = tol_m / 1000.0
    wb = load_workbook(path, data_only=True)
    memos, repairs = [], []
    if "箇所メモ" in wb.sheetnames:
        for row in wb["箇所メモ"].iter_rows(min_row=5, min_col=2, max_col=4,
                                            values_only=True):
            line, km, text = row
            if line and km is not None and text:
                memos.append((str(line), float(km), str(text)))
    memos = list(dict.fromkeys(memos))
    stored = len(memos)

    if "推移表" in wb.sheetnames:
        for row in wb["推移表"].iter_rows(min_row=DATA_START, max_col=COL_MEMO,
                                          values_only=True):
            if not row[0]:
                break
            line, km, text = row[1], row[3], row[COL_MEMO - 1]
            if not (line and km is not None and text):
                continue
            line, km, text = str(line), float(km), str(text)
            # 自動表示された分は「 / 」で連結されている場合があるので分解
            for part in (t.strip() for t in text.split("/")):
                if not part:
                    continue
                known = any(m[0] == line and m[2] == part
                            and abs(m[1] - km) <= tol for m in memos)
                if not known:
                    memos.append((line, km, part))

    if "施工記録" in wb.sheetnames:
        for row in wb["施工記録"].iter_rows(min_row=5, min_col=2, max_col=5,
                                            values_only=True):
            line, km, date, txt = row
            if line and km is not None:
                repairs.append((str(line), float(km), date, txt or ""))
    repairs = list(dict.fromkeys(repairs))
    print(f"引き継ぎ: メモ{len(memos)}件"
          f"(保管{stored} + 推移表から新規{len(memos) - stored}) "
          f"/ 施工記録{len(repairs)}件")
    return memos, repairs


def fill_from_pdfs(wb, specs, tol_m, target, limit, memos=(), repairs=()):
    """specs: [(月番号1-12, PDFパス), ...] を取り込んで推移表・グラフまで作る。"""
    tol = tol_m / 1000.0
    months, metas = {}, {}
    for mi, path in specs:
        meta, rows = parse_pdf(path)
        months.setdefault(mi, []).extend(rows)
        metas[mi] = meta
        print(f"{MONTH_LABELS[mi - 1]}: {path} → {len(rows)}行")

    all_entries, over, tgt_rows, points = [], [], [], []
    for mi in sorted(months):
        label = MONTH_LABELS[mi - 1]
        for row in months[mi]:
            e = (label, metas[mi], row)
            all_entries.append(e)
            v = row["ud"] if row["ud"] is not None else row["sa"]
            if v is None:
                continue
            if v >= limit:
                over.append(e)
            elif v >= target:
                tgt_rows.append(e)
            kind = "上下動" if row["ud"] is not None else "左右動"
            points.append((metas[mi]["line"], kind, row["km"], v, mi,
                           row["biko"]))

    write_data_rows(wb["全データ"], all_entries)
    write_data_rows(wb["①基準値超過"], over)
    write_data_rows(wb["②目標値超過"], tgt_rows)

    ws = wb["取込履歴"]
    for i, mi in enumerate(sorted(months)):
        vals = [MONTH_LABELS[mi - 1], metas[mi]["line"], metas[mi]["date"],
                metas[mi]["file"].split("/")[-1], len(months[mi]), ""]
        for col, v in enumerate(vals, start=2):
            c = ws.cell(row=5 + i, column=col, value=v)
            c.border = BORDER
            c.font = Font(name=FONT, size=10)

    ws = wb["施工記録"]
    for i, (line, km, date, txt) in enumerate(repairs):
        for col, v in enumerate((line, km, date, txt), start=2):
            c = ws.cell(row=5 + i, column=col, value=v)
            c.font = Font(name=FONT, size=10)

    clusters = group_points(points, tol)
    write_memo_sheet(wb, memos)
    write_trend(wb, clusters, sorted(months), tol, repairs, memos)
    write_chart_data(wb, clusters, sorted(months), target, limit,
                     selected=1 if clusters else None)
    print(f"推移表: {len(clusters)}箇所")


def main():
    ap = argparse.ArgumentParser(
        description="動揺データ年度推移管理ブックの生成",
        epilog="例: python build_trend_workbook.py -o 2026年度.xlsx "
               "--pdf 4:4月.pdf --pdf 5:5月.pdf --pdf 6:6月.pdf")
    ap.add_argument("-o", "--output", required=True)
    ap.add_argument("--pdf", action="append", metavar="月:PDF", default=[],
                    help="暦月とPDFの組(例 6:202606.pdf)。複数指定可")
    ap.add_argument("--tol", type=float, default=20,
                    help="同一箇所とみなす距離(m) 既定20")
    ap.add_argument("--target", type=float, default=0.20, help="目標値")
    ap.add_argument("--limit", type=float, default=0.25, help="基準値")
    ap.add_argument("--carry", metavar="既存.xlsx",
                    help="既存ブックの手入力分(メモ・施工記録)を引き継ぐ")
    ap.add_argument("--demo", metavar="PDF",
                    help="実PDFを6月分として疑似データ入りデモを生成")
    args = ap.parse_args()

    wb = build_book()
    if args.demo:
        fill_demo(wb, args.demo)
    elif args.pdf:
        memos, repairs = (read_carry(args.carry, args.tol)
                          if args.carry else ([], []))
        specs = []
        for s in args.pdf:
            month, _, path = s.partition(":")
            specs.append((fiscal_index(int(month)), path))
        fill_from_pdfs(wb, specs, args.tol, args.target, args.limit,
                       memos, repairs)
        wb["設定"]["C6"] = args.tol
        wb["設定"]["C3"] = args.target
        wb["設定"]["C4"] = args.limit
    wb.save(args.output)
    print(f"出力: {args.output}")


if __name__ == "__main__":
    main()
