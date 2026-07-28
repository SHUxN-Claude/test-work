#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""動揺データ抽出ツールのテンプレートExcel(.xlsx)を生成する。

生成後、Excel上でVBAモジュール(excel-tool/vba/*.bas)をインポートして
.xlsm形式で保存すると、PDF取込ボタン付きのツールになる。

使い方:
  python build_template.py [-o 出力.xlsx]
"""
import argparse

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from pdf_to_excel import (FONT_NAME, DATA_START, CF_LAST_ROW,
                          build_workbook)


def add_paste_sheet(wb):
    ws = wb.create_sheet("PDF貼付", index=1)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 14

    t = ws.cell(row=2, column=2, value="■ PDF貼付シート")
    t.font = Font(name=FONT_NAME, bold=True, size=14)

    steps = [
        "【使い方1: ファイル選択(推奨)】",
        "  「PDFファイルを選択して取込」ボタンを押し、配布されたPDFを選ぶだけです(複数選択可)。",
        "",
        "【使い方2: このシートにPDFを貼り付けて取込】",
        "  1. このシートの下の空きスペースに、PDFファイルを貼り付けます。",
        "     ・エクスプローラーからPDFファイルをコピー(Ctrl+C)し、このシートで貼り付け(Ctrl+V)",
        "     ・または「挿入」→「テキスト」→「オブジェクト」→「ファイルから」→「アイコンで表示」",
        "  2. ブックを保存します(貼り付けたPDFはブックの中に保存されます)。",
        "  3. 「貼り付けたPDFを取り込む」ボタンを押すと、貼り付けたPDFの中身を読み取り、",
        "     「全データ」「①基準値超過」「②目標値超過」シートが作成されます。",
        "",
        "【設定を変えたとき】",
        "  「設定」シートの値を変更したら「設定変更後の再抽出」ボタンを押してください。",
        "  (PDFを読み直さずに、取り込み済みのデータから①②を作り直します)",
        "",
        "※ ボタンが表示されていない場合は、Alt+F8 →「初期設定」を実行してください。",
        "※ OneDrive上に保存していると「貼り付けたPDFの取込」が使えない場合があります。",
        "   その場合はローカル(デスクトップ等)に保存するか、使い方1をご利用ください。",
    ]
    for i, s in enumerate(steps):
        c = ws.cell(row=9 + i, column=2, value=s)
        c.font = Font(name=FONT_NAME, size=10,
                      bold=s.startswith("【"))
    note = ws.cell(row=4, column=2,
                   value="ボタンは初回に Alt+F8 →「初期設定」を実行すると設置されます。")
    note.font = Font(name=FONT_NAME, size=10, color="C00000")
    area = ws.cell(row=32, column=2, value="▼ この下にPDFを貼り付けてください ▼")
    area.font = Font(name=FONT_NAME, size=11, bold=True, color="1F4E79")


def preformat_data_columns(wb):
    """VBAが値を書き込むだけで済むよう、データ列の表示形式を先に設定しておく。"""
    fmts = {"B": "0.0", "C": "@", "D": "0.000", "E": "0.00", "F": "0.00"}
    for name in ("全データ", "①基準値超過", "②目標値超過"):
        ws = wb[name]
        for col, fmt in fmts.items():
            for r in range(DATA_START, CF_LAST_ROW + 1):
                ws[f"{col}{r}"].number_format = fmt


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("-o", "--output",
                    default="excel-tool/動揺データ抽出ツール_テンプレート.xlsx")
    args = ap.parse_args()

    build_workbook([], 0.20, 0.25, 0.24, args.output, template_only=True)
    wb = load_workbook(args.output)
    add_paste_sheet(wb)
    preformat_data_columns(wb)
    wb.save(args.output)
    print(f"出力: {args.output}")


if __name__ == "__main__":
    main()
