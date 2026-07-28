#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""動揺データPDF → Excel 変換スクリプト

軌道検測(動揺測定)の結果PDFから「キロ程・上下動・左右動」の表を抽出し、
以下のシートを持つExcelブックを生成する。

  設定        … 目標値・基準値・黄色強調値(職場ごとに変更して共有可能)
  全データ    … PDFの全行
  ①基準値超過 … 上下動もしくは左右動が「基準値」以上の行
  ②目標値超過 … 上下動もしくは左右動が「目標値」以上「基準値」未満の行
                 (「黄色強調値」以上のセルは黄色の条件付き書式で強調)

使い方:
  python pdf_to_excel.py 入力1.pdf [入力2.pdf ...] -o 出力.xlsx
      [--target 0.20] [--limit 0.25] [--yellow 0.24]
"""
import argparse
import re
import sys

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

FONT_NAME = "游ゴシック"

RE_NO = re.compile(r"^\d+$")
RE_TIME = re.compile(r"^\d{1,2}:\d{2}:\d{2}$")
RE_KM = re.compile(r"^\d+\.\d{3}$")
RE_VAL = re.compile(r"^\d\.\d{2}$")


def parse_pdf(path):
    """PDFを解析し (meta, rows) を返す。

    rows の各要素:
      dict(no, speed, time, km, ud, ud_star, sa, sa_star, biko)
    上下動/左右動の列判定は、各ページの「上下動」「左右動」見出しの
    x座標中心と数値のx座標中心の距離で行う。
    """
    meta = {"title": "", "center": "", "line": "", "date": "", "file": path}
    rows = []
    ud_c = sa_c = None
    with pdfplumber.open(path) as pdf:
        first_text = pdf.pages[0].extract_text() or ""
        m = re.search(r"保線技術センター：(\S+)", first_text)
        if m:
            meta["center"] = m.group(1)
        m = re.search(r"線別：(\S+)", first_text)
        if m:
            meta["line"] = m.group(1)
        m = re.search(r"測定日：(\S+)", first_text)
        if m:
            meta["date"] = m.group(1)
        m = re.search(r"表題：\s*(.*?)\s*(?:保線技術センター：|$)", first_text)
        if m:
            meta["title"] = m.group(1)

        for page in pdf.pages:
            words = page.extract_words()
            for w in words:
                if w["text"] == "上下動":
                    ud_c = (w["x0"] + w["x1"]) / 2
                elif w["text"] == "左右動":
                    sa_c = (w["x0"] + w["x1"]) / 2
            if ud_c is None or sa_c is None:
                continue
            lines = {}
            for w in words:
                lines.setdefault(round(w["top"]), []).append(w)
            for key in sorted(lines):
                ws = sorted(lines[key], key=lambda w: w["x0"])
                texts = [w["text"] for w in ws]
                if len(texts) < 4 or not RE_NO.fullmatch(texts[0]):
                    continue
                # 時刻・キロ程はトークンの並び順ではなく書式で特定する
                # (時刻とキロ程の間に「W」等の記号が入る行があるため)
                time_tok = next((t for t in texts if RE_TIME.fullmatch(t)), None)
                km_w = next((w for w in ws if RE_KM.fullmatch(w["text"])), None)
                if time_tok is None or km_w is None:
                    continue
                row = dict(no=int(texts[0]), speed=float(texts[1]),
                           time=time_tok, km=float(km_w["text"]),
                           ud=None, ud_star=False, sa=None, sa_star=False,
                           biko="")
                biko = []
                for w in ws[2:]:
                    t = w["text"]
                    if w is km_w or RE_TIME.fullmatch(t):
                        continue
                    c = (w["x0"] + w["x1"]) / 2
                    if RE_VAL.fullmatch(t) and w["x0"] > km_w["x1"]:
                        if abs(c - ud_c) < abs(c - sa_c):
                            row["ud"] = float(t)
                        else:
                            row["sa"] = float(t)
                    elif t == "*":
                        # 「*」は数値の左隣に付く超過マーク
                        if abs(c - ud_c) < abs(c - sa_c) + 15:
                            row["ud_star"] = True
                        else:
                            row["sa_star"] = True
                    elif w["x0"] > 285:
                        # 左右動の数値域より右側 → 備考欄
                        biko.append(t)
                    elif w["x0"] > 160 and w["x0"] < km_w["x0"]:
                        # キロ程の左に付く記号(例:「W」) → 備考の先頭に付記
                        biko.insert(0, "(" + t + ")")
                row["biko"] = " ".join(biko)
                rows.append(row)
    return meta, rows


HEADERS = ["NO", "列車速度\n(km/h)", "時刻\n(H:M:S)", "キロ程\n(km)",
           "上下動\n(G)", "左右動\n(G)", "超過印", "備考",
           "測定日", "線別", "保線技術センター", "元ファイル"]
COL_WIDTHS = [6, 10, 10, 10, 9, 9, 7, 18, 12, 8, 16, 28]

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", start_color="D9E1F2")
YELLOW_FILL = PatternFill("solid", start_color="FFFF00")
RED_FILL = PatternFill("solid", start_color="FFC7CE")
RED_FONT = Font(name=FONT_NAME, color="9C0006", bold=True)
INPUT_FILL = PatternFill("solid", start_color="FFF2CC")

DATA_START = 5          # データは5行目から
CF_LAST_ROW = 6000      # 条件付き書式を適用しておく最終行


def _setup_data_sheet(ws, note_formula):
    ws.sheet_view.showGridLines = True
    for i, (h, wdt) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name=FONT_NAME, bold=True, size=10)
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.row_dimensions[4].height = 28
    ws.freeze_panes = "A5"
    a2 = ws.cell(row=2, column=1, value=note_formula)
    a2.font = Font(name=FONT_NAME, size=10, color="1F4E79")
    ws.cell(row=2, column=6, value="該当件数：").font = Font(name=FONT_NAME, size=10)
    cnt = ws.cell(row=2, column=7,
                  value=f"=COUNT(E{DATA_START}:F{CF_LAST_ROW})")
    cnt.font = Font(name=FONT_NAME, size=10, bold=True)
    # 条件付き書式(設定シートの名前付きセルを参照するため、設定変更に追従する)
    rng = f"E{DATA_START}:F{CF_LAST_ROW}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f"AND(ISNUMBER(E{DATA_START}),E{DATA_START}>=基準値)"],
        fill=RED_FILL, font=RED_FONT, stopIfTrue=True))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f"AND(ISNUMBER(E{DATA_START}),E{DATA_START}>=黄色強調値,"
                 f"E{DATA_START}<基準値)"],
        fill=YELLOW_FILL, stopIfTrue=True))


def _write_rows(ws, entries):
    r = DATA_START
    for meta, row in entries:
        values = [row["no"], row["speed"], row["time"], row["km"],
                  row["ud"], row["sa"],
                  "*" if (row["ud_star"] or row["sa_star"]) else "",
                  row["biko"], meta["date"], meta["line"], meta["center"],
                  meta["file"].split("/")[-1].split("\\")[-1]]
        fmts = [None, "0.0", "@", "0.000", "0.00", "0.00",
                None, None, None, None, None, None]
        for i, (v, f) in enumerate(zip(values, fmts), start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=FONT_NAME, size=10)
            c.border = BORDER
            if f:
                c.number_format = f
            if i in (1, 3, 7, 9, 10):
                c.alignment = Alignment(horizontal="center")
        r += 1


def build_settings_sheet(ws, target, limit, yellow):
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 70
    t = ws.cell(row=1, column=2, value="■ 設定シート(職場ごとにここだけ変更してください)")
    t.font = Font(name=FONT_NAME, bold=True, size=12)

    items = [
        ("目標値", target, "この値以上を「②目標値超過」シートに抽出します。"),
        ("基準値", limit, "この値以上を「①基準値超過」シートに抽出します。"),
        ("黄色強調値", yellow, "②のシート内で、この値以上のセルを黄色で強調します。"),
    ]
    for i, (name, val, desc) in enumerate(items):
        r = 3 + i
        ws.cell(row=r, column=2, value=name).font = Font(name=FONT_NAME, bold=True)
        c = ws.cell(row=r, column=3, value=val)
        c.font = Font(name=FONT_NAME, color="0000FF", bold=True)
        c.fill = INPUT_FILL
        c.border = BORDER
        c.number_format = "0.00"
        ws.cell(row=r, column=4, value=desc).font = Font(name=FONT_NAME, size=10)

    notes = [
        "",
        "【凡例】 色付き(うすい橙)のセルが入力欄です。数値のみ変更してください。",
        "",
        "【例1】 目標値0.20/基準値0.25の職場 → 目標値=0.20、基準値=0.25、黄色強調値=0.24",
        "【例2】 目標値0.25/基準値0.30の職場 → 目標値=0.25、基準値=0.30、黄色強調値=0.29",
        "",
        "※ 設定を変更したら「再抽出」(マクロ版)を実行するか、PDFを取り込み直してください。",
        "※ 黄色・赤色の強調(条件付き書式)は設定値を参照しているため、自動で追従します。",
    ]
    for i, txt in enumerate(notes):
        ws.cell(row=7 + i, column=2, value=txt).font = Font(name=FONT_NAME, size=10)


def build_workbook(parsed, target, limit, yellow, out_path,
                   template_only=False):
    wb = Workbook()
    ws_set = wb.active
    ws_set.title = "設定"
    build_settings_sheet(ws_set, target, limit, yellow)
    for name, cell in (("目標値", "$C$3"), ("基準値", "$C$4"),
                       ("黄色強調値", "$C$5")):
        wb.defined_names.add(DefinedName(name, attr_text=f"設定!{cell}"))

    ws_all = wb.create_sheet("全データ")
    ws_over = wb.create_sheet("①基準値超過")
    ws_tgt = wb.create_sheet("②目標値超過")

    _setup_data_sheet(ws_all, '="PDFから読み込んだ全データ"')
    _setup_data_sheet(
        ws_over,
        '="抽出条件： 上下動もしくは左右動が 基準値("&TEXT(基準値,"0.00")&"G)以上"')
    _setup_data_sheet(
        ws_tgt,
        '="抽出条件： 目標値("&TEXT(目標値,"0.00")&"G)以上〜基準値("'
        '&TEXT(基準値,"0.00")&"G)未満　※ "&TEXT(黄色強調値,"0.00")'
        '&"G以上は黄色で強調"')

    if not template_only:
        all_entries, over, tgt = [], [], []
        for meta, rows in parsed:
            for row in rows:
                v = row["ud"] if row["ud"] is not None else row["sa"]
                all_entries.append((meta, row))
                if v is None:
                    continue
                if v >= limit:
                    over.append((meta, row))
                elif v >= target:
                    tgt.append((meta, row))
        _write_rows(ws_all, all_entries)
        _write_rows(ws_over, over)
        _write_rows(ws_tgt, tgt)

    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser(description="動揺データPDF → Excel 変換")
    ap.add_argument("pdfs", nargs="+", help="入力PDF(複数可)")
    ap.add_argument("-o", "--output", default="動揺データ抽出結果.xlsx")
    ap.add_argument("--target", type=float, default=0.20, help="目標値")
    ap.add_argument("--limit", type=float, default=0.25, help="基準値")
    ap.add_argument("--yellow", type=float, default=0.24, help="黄色強調値")
    args = ap.parse_args()

    parsed = []
    for p in args.pdfs:
        meta, rows = parse_pdf(p)
        print(f"{p}: {len(rows)}行 (測定日 {meta['date']} / {meta['line']})")
        parsed.append((meta, rows))

    build_workbook(parsed, args.target, args.limit, args.yellow, args.output)
    print(f"出力: {args.output}")


if __name__ == "__main__":
    sys.exit(main())
